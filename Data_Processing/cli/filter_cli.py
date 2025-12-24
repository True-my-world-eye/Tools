import os
import re
import time
from typing import List, Dict, Optional, Tuple, Iterable

"""
跨平台CLI批量筛选（百万行/≤500条件）
依赖：pandas、openpyxl（可选：rapidfuzz用于加速模糊匹配）
用法：直接运行该脚本；参数在代码顶部配置

设计说明（概览）：
- 输入：一个或多个 Excel/CSV 文件，支持 Sheet 多表合并
- 条件：不超过 500 条的 CSV 条件，按列进行向量化评估（尽量避免逐行逐条件的嵌套循环）
- 性能：使用“分块处理”（CHUNK_SIZE）与“向量化操作”，并对文本包含类条件进行合并与预编译
- 输出：逐文件筛选结果与合并结果（可选），支持追加与去重，优先写 Excel，失败降级 CSV
- 进度：终端定期输出处理行数、总计、速率，方便观察运行情况

重要约定：
- 该脚本面向大数据量场景（百万级行），因此尽量减少 Python 层循环，更多使用 pandas 的列操作
- fuzzy（模糊）匹配较耗时，建议结合 code_prefer 精确命中编码，再对剩余进行相似评估；安装 rapidfuzz 可显著提速
"""

# ===================== 配置区域 =====================
EXCEL_FILES: List[str] = [
    # 示例：填写你的数据文件路径（支持多个，绝对/相对路径均可）
    "F:/AAAAclass/python/lzf数据分析/TMT_FIGUREINFO1.xlsx",
]
# SHEET 控制读取的工作表：
# - ""：每个 Excel 文件读取首个工作表
# - "Sheet1,Sheet2"：指定多个工作表并纵向合并
# - "*"：合并该文件所有工作表
SHEET: str = ""
# CONDITIONS_CSV：条件文件路径（CSV/Excel）；为空将回退到“仅Major列”的旧逻辑（不建议）
CONDITIONS_CSV: Optional[str] = "F:/AAAAclass/python/lzf数据分析/combined_conditions_full.csv"

# 旧版（仅专业列）回退配置
MAJOR_COL: str = "Major"        # 旧逻辑使用的专业列名
MAJOR_THRESHOLD: float = 0.80   # 旧逻辑模糊阈值（0~1或百分比，建议提供条件CSV替代旧逻辑）

# 组合与输出
COMBINE_MODE: str = "OR"       # 条件组合模式：AND | OR | WEIGHTED
COMBINE_THRESHOLD: float = 0.80 # WEIGHTED总阈值（0~1）
OUT_DIR: Optional[str] = None   # 逐文件输出目录；None→源目录
MERGE_OUT: Optional[str] = "merged_filtered.xlsx" # 合并输出文件；None→默认 merged_filtered.xlsx（首个输入文件目录）
APPEND: bool = False            # 追加模式：True→读旧结果并合并，False→覆盖
DEDUP: bool = True              # 是否启用去重
DEDUP_KEY: Optional[str] = "PersonID" # 去重键列名；None→回退“规范化Major+编码”（旧逻辑兼容）

# 性能与日志
CHUNK_SIZE: int = 50000            # 分块行数（建议5万~10万；越大内存占用越多）
PROGRESS_STEP: int = 5000         # 每处理N行输出一次进度
WRITE_AUDIT_COLUMNS: bool = False   # 是否写出每条件审计列（便于调试；关闭更轻量）

# ===================== 工具函数 =====================
def to_halfwidth(s: str) -> str:
    """
    将字符串中的全角字符转换为半角，统一符号形态，降低匹配时的格式差异。
    参数：
      s：输入字符串
    返回：
      半角化后的字符串
    """
    r = []
    for ch in s:
        code = ord(ch)
        if code == 0x3000:
            code = 32
        elif 0xFF01 <= code <= 0xFF5E:
            code -= 0xFEE0
        r.append(chr(code))
    return "".join(r)

def normalize_text(s: str) -> str:
    """
    对文本进行规范化处理：
    - 半角化
    - 去首尾空白
    - 转小写
    - 压缩内部空格（保留词界，便于contains与token运算）
    """
    s = to_halfwidth(s)
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)  # 仅压缩空格，保留词界
    return s

def extract_code(s: str) -> str:
    """
    从文本中提取4~6位的数字编码（忽略后缀字母），用于编码优先匹配。
    参数：
      s：原始文本
    返回：
      纯数字编码字符串；未匹配返回空串
    """
    m = re.search(r"(\d{4,6}[A-Z]{0,3})", s or "")
    if not m:
        return ""
    return re.sub(r"[^0-9]", "", m.group(1))

def resolve_path(p: str) -> str:
    """
    将相对路径转换为绝对路径；绝对路径原样返回。
    用于统一文件路径处理，便于跨平台。
    """
    if os.path.isabs(p):
        return p
    return os.path.abspath(p)

def ensure_pandas():
    """
    安全导入 pandas，缺失依赖时给出安装提示。
    返回：
      pandas 模块对象
    """
    try:
        import pandas as pd  # type: ignore
        return pd
    except Exception:
        raise RuntimeError("需要安装依赖：pip install pandas openpyxl")

def read_conditions_csv(pd, path: str) -> List[Dict[str, str]]:
    """
    读取并校验条件CSV/Excel文件：
    - 必须包含表头：column,type,operator,value,threshold,priority,weight,options
    - 将NaN填充为空字符串；统一为字符串字典列表
    """
    path = resolve_path(path)
    if not os.path.exists(path):
        raise FileNotFoundError(f"条件文件不存在：{path}")
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        df = pd.read_excel(path)
    else:
        df = pd.read_csv(path)
    required = ["column","type","operator","value","threshold","priority","weight","options"]
    for c in required:
        if c not in df.columns:
            raise RuntimeError(f"条件文件缺少列：{c}")
    return [{k: str(v).strip() for k, v in r.items()} for r in df.fillna("").to_dict("records")]

def parse_options(s: str) -> Dict[str, str]:
    """
    解析 options 字段（以 ; 分隔的键值对），支持不带=的布尔开关。
    示例：ignore_case=true;normalize=true;code_prefer=true
    """
    opts = {}
    for part in (s or "").split(";"):
        part = part.strip()
        if not part:
            continue
        if "=" in part:
            k, v = part.split("=", 1)
            opts[k.strip()] = v.strip()
        else:
            opts[part] = "true"
    return opts

def build_sheet_frames(pd, excel_path: str, sheet: str) -> List[Tuple[str, Optional[str]]]:
    """
    生成“文件-工作表”组合列表，用于后续分块读取。
    规则：
      - ""：返回 (文件, None) → 读取首个工作表
      - "*"：返回 (文件, "*") → 读取所有工作表
      - "A,B"：返回 (文件,"A"),(文件,"B")
    """
    # 返回 (文件路径, 工作表名或None表示首个) 列表
    excel_path = resolve_path(excel_path)
    if not sheet or sheet.strip() == "":
        return [(excel_path, None)]
    s = sheet.strip()
    if s == "*":
        # 标识所有工作表，实际读取时获取列表
        return [(excel_path, "*")]
    names = [x.strip() for x in s.split(",") if x.strip()]
    return [(excel_path, nm) for nm in names]

def chunk_generator_from_excel(pd, excel_path: str, sheet: Optional[str], chunk_size: int) -> Iterable:
    """
    Excel 流式分块读取：
    - 使用 openpyxl 的 read_only 模式按行读取，避免一次性将整个表加载到内存
    - 每读满 chunk_size 行就产出一个 DataFrame 块
    - 自动处理标题行（首行）为列名
    """
    # Excel流式读取：openpyxl逐行→DataFrame分块
    from openpyxl import load_workbook  # type: ignore
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames if sheet == "*" else ([sheet] if sheet else [wb.sheetnames[0]])
    for nm in sheet_names:
        if nm not in wb.sheetnames:
            print(f"警告：{os.path.basename(excel_path)} 缺少工作表 {nm}，已跳过")
            continue
        ws = wb[nm]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            continue
        header = [str(h) if h is not None else "" for h in header]
        buf = []
        for row in rows_iter:
            buf.append({header[i]: row[i] for i in range(len(header))})
            if len(buf) >= chunk_size:
                yield pd.DataFrame(buf)
                buf.clear()
        if buf:
            yield pd.DataFrame(buf)
    wb.close()

def chunk_generator_from_csv(pd, csv_path: str, chunk_size: int) -> Iterable:
    """
    CSV 分块读取：
    - 直接使用 pandas.read_csv(chunksize=...) 迭代返回 DataFrame块
    """
    yield from pd.read_csv(csv_path, chunksize=chunk_size)

def total_rows_excel(excel_path: str, sheet: Optional[str]) -> int:
    """
    估算Excel总行数（不含标题行），用于计算已处理占比：
    - 使用 openpyxl 的 max_row 快速获取每工作表的行数
    - 当指定多个工作表或 '*' 时求和
    - 若工作表不存在则跳过并提示
    """
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        if sheet == "*":
            names = wb.sheetnames
        elif sheet:
            names = [sheet]
        else:
            names = [wb.sheetnames[0]] if wb.sheetnames else []
        total = 0
        for nm in names:
            if nm not in wb.sheetnames:
                print(f"警告：{os.path.basename(excel_path)} 缺少工作表 {nm}，占比计算跳过该表")
                continue
            ws = wb[nm]
            rows = max(ws.max_row - 1, 0)  # 扣除标题行
            total += rows
        wb.close()
        return total
    except Exception:
        return 0

def total_rows_csv(csv_path: str) -> int:
    """
    估算CSV总行数（不含标题行）：统计文件行数-1
    """
    try:
        with open(csv_path, "r", encoding="utf-8", errors="ignore") as f:
            cnt = sum(1 for _ in f)
        return max(cnt - 1, 0)
    except Exception:
        return 0

def format_time(secs: float) -> str:
    """
    将秒数格式化为 HH:MM:SS
    """
    secs = int(secs)
    h = secs // 3600
    m = (secs % 3600) // 60
    s = secs % 60
    return f"{h:02d}:{m:02d}:{s:02d}"

def render_progress(done: int, total: int, width: int = 30) -> str:
    """
    渲染文本进度条：
    - 当 total=0 时显示未知占比（空进度条）
    - 样式：[######--------------] 23%
    """
    if total <= 0 or done < 0:
        bar = "-" * width
        return f"[{bar}] ??%"
    pct = max(0, min(100, int(done * 100 / total)))
    filled = int(width * pct / 100)
    bar = "#" * filled + "-" * (width - filled)
    return f"[{bar}] {pct:02d}%"

def compile_text_operations(pd, df, conditions: List[Dict[str, str]]) -> Dict:
    """
    对“text contains”类条件按列进行合并与预编译：
    - 将同列、同选项（如 ignore_case）的多个词合并为一个大regex，提升匹配性能
    - 返回：列名→预编译regex
    """
    # 将相同列/相同选项的 contains 合并为一个大regex，提高效率
    col_ops = {}
    for cond in conditions:
        col = cond["column"]
        typ = cond["type"]
        op = cond["operator"]
        opts = cond.get("options","")
        if typ == "text" and op == "contains":
            key = (col, "contains", opts)
            col_ops.setdefault(key, []).append(cond["value"])
    compiled = {}
    import re as _re
    for (col, _, opts), tokens in col_ops.items():
        tokens = [t for t in tokens if t]
        if not tokens:
            continue
        ignore_case = parse_options(opts).get("ignore_case","false").lower()=="true"
        pat = "|".join([_re.escape(t) for t in tokens])
        flags = _re.IGNORECASE if ignore_case else 0
        compiled[col] = _re.compile(pat, flags)
    return compiled

def eval_conditions_block(pd, df, conditions: List[Dict[str, str]], combine_mode: str, combine_threshold: float, write_audit: bool) -> Tuple:
    """
    对一个数据块（DataFrame）执行条件评估（向量化）：
    - 预编译 text contains 与 regex
    - number/enum/boolean/code：广播比较或集合匹配
    - fuzzy：优先 rapidfuzz；否则退化为 normalize+contains（相对更快）
    - 组合：AND/OR/WEIGHTED，生成 _match_all 与 _score_all
    - 审计列：可选输出每条件的命中与分数与描述，便于回溯
    返回：
      (更新后的df, 审计列名列表)
    """
    # 预编译contains大regex
    contains_compiled = compile_text_operations(pd, df, conditions)
    # 预编译regex匹配
    regex_compiled = {}
    for cond in conditions:
        if cond["type"]=="regex" and cond["operator"]=="match" and cond["value"]:
            try:
                regex_compiled[id(cond)] = re.compile(cond["value"])
            except Exception:
                regex_compiled[id(cond)] = None
    # 抽取编码列（如有）
    code_cache = {}
    def get_code_series(column: str):
        s = code_cache.get(column)
        if s is None:
            s = df[column].astype(str).fillna("").map(extract_code) if column in df.columns else pd.Series([""]*len(df))
            code_cache[column] = s
        return s
    # 分数与命中
    total_score = pd.Series([0.0]*len(df))
    any_hit = pd.Series([False]*len(df))
    all_hit = pd.Series([True]*len(df))
    # 审计列容器
    audit_cols = []
    # 向量化评估
    for idx, cond in enumerate(conditions, start=1):
        col = cond["column"]
        typ = cond["type"]
        op = cond["operator"]
        val = cond["value"]
        th_raw = cond.get("threshold","")
        w = float(cond.get("weight","1") or "1")
        opts = parse_options(cond.get("options",""))
        series = df[col].astype(str).fillna("") if col in df.columns else pd.Series([""]*len(df))
        hit = pd.Series([False]*len(df))
        score = pd.Series([0.0]*len(df))
        try:
            if typ == "text":
                if op == "equals":
                    target = val if not opts.get("ignore_case","").lower()=="true" else val.lower()
                    scomp = series if not opts.get("ignore_case","").lower()=="true" else series.str.lower()
                    hit = (scomp == target)
                    score = hit.astype(float)
                elif op == "contains":
                    creg = contains_compiled.get(col)
                    if creg:
                        hit = series.str.contains(creg, na=False)
                        score = hit.astype(float)
                    else:
                        hit = series.str.contains(re.escape(val), case=not opts.get("ignore_case","").lower()=="true", na=False)
                        score = hit.astype(float)
                elif op == "startswith":
                    scomp = series if not opts.get("ignore_case","").lower()=="true" else series.str.lower()
                    target = val if not opts.get("ignore_case","").lower()=="true" else val.lower()
                    hit = scomp.str.startswith(target)
                    score = hit.astype(float)
                elif op == "endswith":
                    scomp = series if not opts.get("ignore_case","").lower()=="true" else series.str.lower()
                    target = val if not opts.get("ignore_case","").lower()=="true" else val.lower()
                    hit = scomp.str.endswith(target)
                    score = hit.astype(float)
            elif typ == "enum" and op == "in":
                items = [x.strip() for x in val.split(";") if x.strip()]
                hit = series.isin(items)
                score = hit.astype(float)
            elif typ == "number":
                s_num = pd.to_numeric(series, errors="coerce")
                if op == "between":
                    parts = val.replace(" ","").split("-")
                    lo = float(parts[0]); hi = float(parts[1])
                    hit = (s_num >= lo) & (s_num <= hi)
                    score = hit.astype(float)
                elif op == "min":
                    lo = float(val); hit = s_num >= lo; score = hit.astype(float)
                elif op == "max":
                    hi = float(val); hit = s_num <= hi; score = hit.astype(float)
                elif op == "equals":
                    eq = float(val); hit = s_num == eq; score = hit.astype(float)
            elif typ == "boolean" and op == "is":
                truth = val.lower() in ("true","1","yes","y","t")
                s_bool = series.str.lower().isin(["true","1","yes","y","t"])
                hit = (s_bool == truth)
                score = hit.astype(float)
            elif typ == "regex" and op == "match":
                creg = regex_compiled.get(id(cond))
                if creg:
                    hit = series.str.contains(creg, na=False)
                    score = hit.astype(float)
            elif typ == "code" and op == "equals":
                s_code = get_code_series(col)
                tgt_code = re.sub(r"[^0-9]","", val)
                hit = (s_code == tgt_code)
                score = hit.astype(float)
            elif typ == "fuzzy" and op == "similar":
                th = 0.0
                if th_raw:
                    th = float(th_raw[:-1])/100.0 if th_raw.endswith("%") else float(th_raw)
                # 优先编码
                if opts.get("code_prefer","").lower()=="true":
                    s_code = get_code_series(col)
                    tgt_code = extract_code(val)
                    hit_code = (s_code == tgt_code) & (tgt_code!="")
                    score = hit_code.astype(float)
                    hit = hit_code
                    # 未命中编码再走相似
                    need_sim = (~hit_code)
                else:
                    need_sim = pd.Series([True]*len(df))
                # 相似：优先使用rapidfuzz，否则退化为normalize+contains
                try:
                    from rapidfuzz import fuzz  # type: ignore
                    s_norm = series.map(normalize_text)
                    tgt_norm = normalize_text(val)
                    sim = s_norm.map(lambda x: (fuzz.token_set_ratio(x, tgt_norm)/100.0))
                    hit_sim = sim >= th
                    score = (score.where(~need_sim, sim)).fillna(sim)
                    hit = hit | hit_sim
                except Exception:
                    s_norm = series.map(normalize_text)
                    tgt_norm = normalize_text(val)
                    hit_sim = s_norm.str.contains(re.escape(tgt_norm), na=False)
                    score = (score.where(~need_sim, hit_sim.astype(float))).fillna(hit_sim.astype(float))
                    hit = hit | hit_sim
            # 组合
            any_hit = any_hit | hit
            all_hit = all_hit & hit
            total_score = total_score + (score * w)
            if write_audit:
                df[f"_cond_{idx}_match"] = hit
                df[f"_cond_{idx}_score"] = score.round(4)
                df[f"_cond_{idx}_desc"] = f"{col}:{typ}/{op}={val}"
                audit_cols.extend([f"_cond_{idx}_match", f"_cond_{idx}_score", f"_cond_{idx}_desc"])
        except Exception as e:
            print(f"条件评估错误（跳过）：{col}:{typ}/{op} -> {e}")
    # 合成总命中
    if combine_mode == "AND":
        match_all = all_hit
    elif combine_mode == "OR":
        match_all = any_hit
    else:
        match_all = total_score >= COMBINE_THRESHOLD
    df["_match_all"] = match_all
    df["_score_all"] = total_score.round(4)
    return df, audit_cols

def write_output(pd, df, out_path: str, append: bool, dedup: bool, dedup_key: Optional[str], major_col: str) -> str:
    """
    写出结果（逐文件或合并）：
    - 优先写 Excel（xlsx/xls），失败降级 CSV（utf-8-sig）
    - 去重：
      · 指定 DEDUP_KEY 且存在：按该列去重
      · 否则回退“规范化 Major + 编码”组合键（旧逻辑兼容）
    - 追加：
      · APPEND=True 且文件存在：读取旧结果，与新结果合并后写出
    返回：
      最终写出的文件路径
    """
    out_path = resolve_path(out_path)
    ext = os.path.splitext(out_path)[1].lower()
    try:
        if dedup:
            if dedup_key and dedup_key in df.columns:
                df = df.drop_duplicates(subset=[dedup_key]).copy()
            else:
                key_series = df[major_col].astype(str).fillna("").map(normalize_text) if major_col in df.columns else pd.Series([""]*len(df))
                code_series = df.get("_matched_code", pd.Series([""]*len(df)))
                df["_dedup_key"] = key_series + "|" + code_series.astype(str)
                df = df.drop_duplicates(subset=["_dedup_key"]).copy()
                df.drop(columns=["_dedup_key"], inplace=True, errors="ignore")
        if append and os.path.exists(out_path):
            if ext in (".xlsx",".xls"):
                old = pd.read_excel(out_path)
            else:
                old = pd.read_csv(out_path)
            df = pd.concat([old, df], ignore_index=True)
        if ext in (".xlsx",".xls"):
            df.to_excel(out_path, index=False)
        else:
            df.to_csv(out_path, index=False, encoding="utf-8-sig")
        return out_path
    except Exception:
        csv_path = os.path.splitext(out_path)[0] + ".csv"
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        return csv_path

def process_files():
    """
    主流程：
    - 读取条件（CSV），为空则回退旧版（仅 Major）逻辑
    - 遍历输入文件：
      · 构造（文件, 工作表）帧列表
      · 按帧分块读取（Excel/CSV）
      · 对块执行条件评估并收集命中结果
    - 写出逐文件结果与（可选）全量合并结果
    - 输出总计处理行数与耗时
    """
    pd = ensure_pandas()
    # 条件加载或回退
    conditions = []
    if CONDITIONS_CSV:
        try:
            conditions = read_conditions_csv(pd, CONDITIONS_CSV)
            print(f"已加载条件 {len(conditions)} 条")
        except Exception as e:
            print(f"条件文件读取失败：{e}")
            conditions = []
    # 旧版回退标记
    use_major_only = (len(conditions) == 0)
    total_written = []
    merged_parts = []
    t0 = time.time()
    total_rows = 0
    for pth in EXCEL_FILES:
        pth = resolve_path(pth)
        if not os.path.exists(pth):
            print(f"文件不存在：{pth}（跳过）")
            continue
        frames = build_sheet_frames(pd, pth, SHEET)
        out_dir = OUT_DIR or os.path.dirname(pth)
        base = os.path.splitext(os.path.basename(pth))[0]
        out_path = os.path.join(out_dir, f"{base}_filtered.xlsx")
        print(f"开始处理：{os.path.basename(pth)}")
        written_this = []
        processed_rows = 0
        file_total_rows = 0
        file_start = time.time()
        file_matched_rows = 0
        # 预估总行数（用于进度占比）
        try:
            if pth.lower().endswith(".csv"):
                file_total_rows = total_rows_csv(pth)
            else:
                # 多帧时占比计算按全部帧求和
                if SHEET == "*" or (SHEET and "," not in SHEET and SHEET.strip() != ""):
                    file_total_rows = total_rows_excel(pth, SHEET if SHEET else None)
                else:
                    # 多工作表名：逐名求和
                    sum_rows = 0
                    names = [x.strip() for x in SHEET.split(",") if x.strip()] if SHEET else [None]
                    for nm in names:
                        sum_rows += total_rows_excel(pth, nm)
                    file_total_rows = sum_rows
        except Exception:
            file_total_rows = 0
        for fp, sh in frames:
            # 分块读取
            if fp.lower().endswith(".csv"):
                chunks = chunk_generator_from_csv(pd, fp, CHUNK_SIZE)
            else:
                chunks = chunk_generator_from_excel(pd, fp, sh, CHUNK_SIZE)
            for block in chunks:
                processed_rows += len(block)
                total_rows += len(block)
                if use_major_only:
                    # 旧版：仅Major列（向量化）
                    s_major = block[MAJOR_COL].astype(str).fillna("") if MAJOR_COL in block.columns else pd.Series([""]*len(block))
                    s_norm = s_major.map(normalize_text)
                    # 简化近似：直接按阈值做normalize+contains（可调整为编码优先）
                    target_norm = ""  # 无具体目标，这里留空 -> 不筛选；旧版需基于require.txt才能生效
                    block["_match_all"] = s_norm.str.len() > 0  # 占位：如需旧版，建议提供CONDITIONS_CSV
                    block["_score_all"] = 1.0
                    audit_cols = []
                else:
                    block, audit_cols = eval_conditions_block(pd, block, conditions, COMBINE_MODE, COMBINE_THRESHOLD, WRITE_AUDIT_COLUMNS)
                out_df = block[block["_match_all"]==True].copy()
                if len(out_df) > 0:
                    written_this.append(out_df)
                    file_matched_rows += len(out_df)
                if PROGRESS_STEP and processed_rows % PROGRESS_STEP == 0:
                    elapsed_file = time.time() - file_start
                    bar = render_progress(processed_rows, file_total_rows)
                    print(f"{bar} 已处理 {processed_rows}/{file_total_rows if file_total_rows>0 else '?'} 行 | 已运行 {format_time(elapsed_file)} | 命中 {file_matched_rows} 行")
        # 写出当前文件结果
        if written_this:
            df_all = pd.concat(written_this, ignore_index=True)
            saved = write_output(pd, df_all, out_path, APPEND, DEDUP, DEDUP_KEY, MAJOR_COL)
            print(f"已写出：{saved}（{len(df_all)} 行）")
            total_written.append(saved)
            merged_parts.append(df_all)
        else:
            print("无命中结果，跳过写出")
    # 合并写出
    if merged_parts:
        df_merged = pd.concat(merged_parts, ignore_index=True)
        m_out = MERGE_OUT or os.path.join(os.path.dirname(EXCEL_FILES[0]) if EXCEL_FILES else os.getcwd(), "merged_filtered.xlsx")
        saved = write_output(pd, df_merged, m_out, APPEND, DEDUP, DEDUP_KEY, MAJOR_COL)
        print(f"合并写出：{saved}（{len(df_merged)} 行）")
    t1 = time.time()
    print(f"完成：总计处理 {total_rows} 行，耗时 {int(t1-t0)} 秒")

if __name__ == "__main__":
    process_files()
